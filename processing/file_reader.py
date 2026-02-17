"""
Excel file reader with automatic structure detection.

Reads raw shelf analysis Excel files and extracts structured data regardless
of layout variation. Handles three known structures:
  A) Flat — header in row 1, data starts in column A, no merged cells.
  B) Sectioned — merged-cell separators divide the sheet into sections,
     each with its own repeated header row. Data starts in column A.
  C) Offset + Sectioned — same as B but data starts in column F or G,
     with a legend/key block in the leftmost columns.

Also detects "context-only rows" (photo/location metadata with no SKU data)
and carries their metadata forward to subsequent data rows, identical to how
merged section separators work.

Public API:
    read_excel_file(file_path) → FileReadResult

See docs/ARCHITECTURE.md for module responsibilities.
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from config.column_mapping import KNOWN_HEADER_NAMES

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Minimum number of non-empty cells for a row to be considered data.
# Rows with fewer populated cells are skipped (catches empty rows and
# single-cell annotation rows like "OTHER TO GO DRINKS").
# ---------------------------------------------------------------------------
MIN_CELLS_FOR_DATA_ROW: int = 3

# Minimum fraction of a row's non-empty cells that must match known header
# names for the row to be classified as a header.
HEADER_MATCH_THRESHOLD: int = 5

# Column names that indicate SKU-level product data (not just section context).
# If a row has none of these populated, it's a context row, not a data row.
_SKU_INDICATOR_COLUMNS: set[str] = {"brand", "flavor", "facings", "segment",
                                     "sub-segment", "product type"}


# ═══════════════════════════════════════════════════════════════════════════
# Data classes
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class SectionMetadata:
    """Metadata extracted from a merged-cell separator or context-only row."""

    photo: str | None = None
    shelf_location: str | None = None
    est_linear_meters: float | None = None
    shelf_levels: int | None = None
    raw_text: str = ""
    start_row: int = 0   # 1-based Excel row where section data begins
    end_row: int = 0     # 1-based Excel row where section data ends


@dataclass
class FileReadResult:
    """Complete result of reading one Excel file."""

    raw_dataframe: pd.DataFrame = field(default_factory=pd.DataFrame)
    sections: list[SectionMetadata] = field(default_factory=list)
    header_row_index: int = 0            # 1-based Excel row of the first header
    data_start_column: int = 0           # 0-based column offset (0=A, 5=F, 6=G)
    sheet_name: str = ""
    total_rows_read: int = 0
    skipped_rows: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════

def read_excel_file(file_path: Path) -> FileReadResult:
    """
    Read a raw shelf analysis Excel file and extract structured data.

    Handles all known file structures: flat, sectioned with merged-cell
    separators, and offset-column layouts.  Detects header rows, skips
    non-data rows, parses section metadata, and returns a clean DataFrame.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        FileReadResult containing the raw DataFrame, section metadata,
        structural info, and any errors encountered.
    """
    result = FileReadResult()

    # ------------------------------------------------------------------
    # 1. Open workbook
    # ------------------------------------------------------------------
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        error_message = f"Cannot open file '{file_path.name}': {exc}"
        logger.error(error_message)
        result.errors.append(error_message)
        return result

    # ------------------------------------------------------------------
    # 2. Select the right sheet (prefer "SKU Data", fall back to first)
    # ------------------------------------------------------------------
    worksheet, sheet_name = _select_worksheet(workbook)
    result.sheet_name = sheet_name
    logger.info(f"Reading sheet '{sheet_name}' from '{file_path.name}'")

    # ------------------------------------------------------------------
    # 3. Detect merged-cell section separators
    # ------------------------------------------------------------------
    merged_separator_rows = _detect_merged_separators(worksheet)
    merged_row_numbers: set[int] = {row_num for row_num, _ in merged_separator_rows}

    # ------------------------------------------------------------------
    # 4. Find the first header row (and data start column)
    # ------------------------------------------------------------------
    header_row, data_start_col = _find_header_row(
        worksheet, KNOWN_HEADER_NAMES, merged_row_numbers
    )

    if header_row == 0:
        error_message = (
            f"No header row found in '{file_path.name}'. "
            "Could not detect columns matching the known schema."
        )
        logger.error(error_message)
        result.errors.append(error_message)
        workbook.close()
        return result

    result.header_row_index = header_row
    result.data_start_column = data_start_col
    logger.info(
        f"Header found at row {header_row}, data starts at column "
        f"{'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[data_start_col]} (offset={data_start_col})"
    )

    # ------------------------------------------------------------------
    # 5. Read column names from the first header row
    # ------------------------------------------------------------------
    column_names = _read_header_columns(worksheet, header_row, data_start_col)

    # ------------------------------------------------------------------
    # 6. Parse section metadata from merged separators
    # ------------------------------------------------------------------
    sections = _build_section_list(
        merged_separator_rows, worksheet.max_row
    )

    # ------------------------------------------------------------------
    # 7. Walk all rows: classify, extract data, carry forward metadata
    # ------------------------------------------------------------------
    raw_dataframe, skipped_rows, context_sections = _extract_rows_to_dataframe(
        worksheet=worksheet,
        column_names=column_names,
        header_row=header_row,
        data_start_col=data_start_col,
        merged_row_numbers=merged_row_numbers,
        sections=sections,
    )

    # Combine merged-cell sections with context-row sections
    all_sections = sections + context_sections
    all_sections.sort(key=lambda s: s.start_row)

    result.raw_dataframe = raw_dataframe
    result.sections = all_sections
    result.total_rows_read = len(raw_dataframe)
    result.skipped_rows = skipped_rows

    logger.info(
        f"Finished reading '{file_path.name}': {result.total_rows_read} data rows, "
        f"{len(all_sections)} sections, {len(skipped_rows)} rows skipped"
    )

    workbook.close()
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Internal helpers
# ═══════════════════════════════════════════════════════════════════════════

def _select_worksheet(
    workbook: openpyxl.Workbook,
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, str]:
    """
    Pick the worksheet to read.

    Prefers a sheet named "SKU Data" (case-insensitive). Falls back to the
    first sheet in the workbook.

    Returns:
        Tuple of (worksheet, sheet_name).
    """
    for name in workbook.sheetnames:
        if name.lower() == "sku data":
            return workbook[name], name

    first_name = workbook.sheetnames[0]
    logger.info(
        f"No 'SKU Data' sheet found; using first sheet '{first_name}'"
    )
    return workbook[first_name], first_name


# ── Merged separator detection ─────────────────────────────────────────

def _detect_merged_separators(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> list[tuple[int, str]]:
    """
    Find all rows that are wide merged cells acting as section separators.

    A merged range qualifies as a separator if it spans at least 10 columns
    (these separators stretch across the entire data width, typically 23+ cols).

    Returns:
        Sorted list of (row_number, cell_text) for each separator.
    """
    separators: list[tuple[int, str]] = []

    for merged_range in worksheet.merged_cells.ranges:
        column_span = merged_range.max_col - merged_range.min_col + 1
        row_span = merged_range.max_row - merged_range.min_row + 1

        # Section separators span many columns but only one row
        if column_span >= 10 and row_span == 1:
            row_number = merged_range.min_row
            # openpyxl stores the value only in the top-left cell
            cell_value = worksheet.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
            text = str(cell_value).strip() if cell_value is not None else ""
            separators.append((row_number, text))

    separators.sort(key=lambda pair: pair[0])
    return separators


def _parse_section_text(raw_text: str) -> SectionMetadata:
    """
    Extract metadata from a merged section separator's text content.

    Handles pipe-delimited format with key:value pairs, for example:
      '📷 Chilled_section_Fridge1 | Location: Chilled Section - Fridge 1 |
       Est. Linear Meters: 2.5 | Shelf Levels: 6'

    Also handles short references like 'Photo: IMG_0118.jpeg' or
    'Photo: Chilled_Section_Fridge1_Top_shelf.jpg  |  Location: Chilled Se...'

    Args:
        raw_text: The raw text from the merged cell.

    Returns:
        SectionMetadata with whatever fields could be parsed.
    """
    metadata = SectionMetadata(raw_text=raw_text)

    if not raw_text:
        return metadata

    # Split on pipe delimiter and process each segment
    segments = [seg.strip() for seg in raw_text.split("|")]

    for segment in segments:
        _parse_segment_into_metadata(segment, metadata)

    # If no explicit Photo key was found, the first segment may be the photo
    # reference (e.g. "📷 Chilled_section_Fridge1" or bare "Photo: X.jpeg")
    if metadata.photo is None and segments:
        first_segment = segments[0].strip()
        # Strip camera emoji and leading whitespace
        cleaned = re.sub(r"^[📷\U0001f4f7\s]+", "", first_segment).strip()
        if cleaned and ":" not in cleaned:
            # Bare photo reference without a key label
            metadata.photo = cleaned

    return metadata


def _parse_segment_into_metadata(segment: str, metadata: SectionMetadata) -> None:
    """
    Parse a single pipe-delimited segment and store any recognized
    key:value pair into the metadata object.

    Args:
        segment: One segment from splitting the separator text on '|'.
        metadata: SectionMetadata to populate (mutated in-place).
    """
    segment_lower = segment.lower().strip()

    # Try to split on the first colon
    if ":" not in segment:
        return

    key_part, _, value_part = segment.partition(":")
    key = key_part.strip().lower()
    value = value_part.strip()

    if not value:
        return

    if key in ("photo", "photo file name"):
        metadata.photo = value
    elif key == "location":
        metadata.shelf_location = value
    elif key in ("est. linear meters", "est linear meters", "linear meters"):
        try:
            metadata.est_linear_meters = float(value)
        except ValueError:
            logger.debug(f"Could not parse linear meters from '{value}'")
    elif key in ("shelf levels", "levels"):
        try:
            metadata.shelf_levels = int(float(value))
        except ValueError:
            logger.debug(f"Could not parse shelf levels from '{value}'")


def _build_section_list(
    merged_separator_rows: list[tuple[int, str]],
    max_row: int,
) -> list[SectionMetadata]:
    """
    Parse each merged separator and assign row ranges to sections.

    Each section starts at the row after its separator (typically the header
    row) and ends at the row before the next separator (or end of sheet).

    Args:
        merged_separator_rows: Sorted list of (row_number, text) pairs.
        max_row: Last row in the worksheet.

    Returns:
        List of SectionMetadata objects with row ranges filled in.
    """
    sections: list[SectionMetadata] = []

    for idx, (row_num, text) in enumerate(merged_separator_rows):
        section = _parse_section_text(text)
        section.start_row = row_num + 1  # Data begins after the separator

        if idx + 1 < len(merged_separator_rows):
            next_separator_row = merged_separator_rows[idx + 1][0]
            section.end_row = next_separator_row - 1
        else:
            section.end_row = max_row

        sections.append(section)

    return sections


# ── Header detection ───────────────────────────────────────────────────

def _find_header_row(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    known_headers: set[str],
    merged_rows: set[int],
) -> tuple[int, int]:
    """
    Scan the worksheet top-down to find the first header row.

    A row qualifies as a header if at least HEADER_MATCH_THRESHOLD of its
    non-empty cell values (lowercased, stripped) appear in *known_headers*.

    Args:
        worksheet: The openpyxl worksheet to scan.
        known_headers: Set of recognised header names (all lowercase).
        merged_rows: Row numbers that are merged separators (skip these).

    Returns:
        (row_number, start_column) — both 1-based.  Returns (0, 0) if no
        header row is found.
    """
    # Only scan the first 30 rows — headers are always near the top
    scan_limit = min(30, worksheet.max_row)

    for row_idx in range(1, scan_limit + 1):
        if row_idx in merged_rows:
            continue

        row_cells = list(
            worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False)
        )[0]

        matches = 0
        first_match_col = 0

        for cell in row_cells:
            if cell.value is None:
                continue
            cell_text = str(cell.value).strip().lower()
            if cell_text in known_headers:
                matches += 1
                if first_match_col == 0:
                    # column_letter is 1-based in openpyxl; .column is 1-based int
                    first_match_col = cell.column

        if matches >= HEADER_MATCH_THRESHOLD:
            # Convert 1-based column to 0-based offset
            start_col_zero_based = first_match_col - 1
            return row_idx, start_col_zero_based

    return 0, 0


def _read_header_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    data_start_col: int,
) -> list[str]:
    """
    Read column names from the header row, starting at *data_start_col*.

    Returns:
        List of column name strings.  Empty cells get placeholder names
        like '_unnamed_0', '_unnamed_1', etc.
    """
    row_cells = list(
        worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=False)
    )[0]

    column_names: list[str] = []
    unnamed_counter = 0

    for cell in row_cells:
        # Skip cells before the data start column
        if cell.column - 1 < data_start_col:
            continue

        if cell.value is not None and str(cell.value).strip():
            column_names.append(str(cell.value).strip())
        else:
            column_names.append(f"_unnamed_{unnamed_counter}")
            unnamed_counter += 1

    return column_names


def _is_header_row(
    row_values: list,
    known_headers: set[str],
    threshold: int = HEADER_MATCH_THRESHOLD,
) -> bool:
    """
    Check whether a row is a repeated header (same column names as first header).

    Args:
        row_values: List of cell values for the row.
        known_headers: Set of recognised header names (lowercase).
        threshold: Minimum matches to classify as a header.

    Returns:
        True if the row looks like a header row.
    """
    matches = 0
    for value in row_values:
        if value is None:
            continue
        text = str(value).strip().lower()
        if text in known_headers:
            matches += 1
    return matches >= threshold


def _is_context_only_row(
    row_values: list,
    column_names: list[str],
) -> bool:
    """
    Detect a "context-only" row that provides section metadata but no SKU data.

    A context row has Photo/Location/Linear Meters populated but Brand,
    Flavor, and Facings are ALL empty.

    Args:
        row_values: Cell values aligned with *column_names*.
        column_names: Column names from the header.

    Returns:
        True if the row is context-only (should be treated as section metadata).
    """
    # Build a quick lookup: lowercased column name → value
    col_value_map: dict[str, object] = {}
    for col_name, value in zip(column_names, row_values):
        col_value_map[col_name.strip().lower()] = value

    # Check if any SKU-indicator column has a non-empty value
    has_sku_data = False
    for indicator in _SKU_INDICATOR_COLUMNS:
        val = col_value_map.get(indicator)
        if val is not None and str(val).strip() != "":
            has_sku_data = True
            break

    if has_sku_data:
        return False

    # Check that the row has at least some context info (photo or location)
    has_context = False
    context_columns = {"photo file name", "photo", "shelf location",
                       "est. linear meters", "shelf levels"}
    for ctx_col in context_columns:
        val = col_value_map.get(ctx_col)
        if val is not None and str(val).strip() != "":
            has_context = True
            break

    return has_context


def _extract_context_metadata(
    row_values: list,
    column_names: list[str],
    row_index: int,
) -> SectionMetadata:
    """
    Build a SectionMetadata from a context-only row.

    Args:
        row_values: Cell values aligned with *column_names*.
        column_names: Column names from the header.
        row_index: 1-based Excel row number.

    Returns:
        SectionMetadata populated from the context row's cells.
    """
    col_value_map: dict[str, object] = {}
    for col_name, value in zip(column_names, row_values):
        col_value_map[col_name.strip().lower()] = value

    metadata = SectionMetadata(
        raw_text=f"Context row at Excel row {row_index}",
        start_row=row_index,
    )

    # Photo
    for key in ("photo file name", "photo"):
        val = col_value_map.get(key)
        if val is not None and str(val).strip():
            metadata.photo = str(val).strip()
            break

    # Shelf Location
    val = col_value_map.get("shelf location")
    if val is not None and str(val).strip():
        metadata.shelf_location = str(val).strip()

    # Est. Linear Meters
    val = col_value_map.get("est. linear meters")
    if val is not None:
        try:
            metadata.est_linear_meters = float(val)
        except (ValueError, TypeError):
            pass

    # Shelf Levels
    val = col_value_map.get("shelf levels")
    if val is not None:
        try:
            metadata.shelf_levels = int(float(val))
        except (ValueError, TypeError):
            pass

    return metadata


def _get_section_for_row(
    row_index: int,
    sections: list[SectionMetadata],
) -> SectionMetadata | None:
    """
    Find the section that a given row belongs to.

    Sections are ordered by start_row. A row belongs to the last section
    whose start_row is <= the row index.

    Args:
        row_index: 1-based Excel row number.
        sections: List of sections sorted by start_row.

    Returns:
        The matching SectionMetadata, or None if no section covers this row.
    """
    best: SectionMetadata | None = None
    for section in sections:
        if section.start_row <= row_index:
            best = section
        else:
            break
    return best


def _apply_section_defaults(
    row_dict: dict[str, object],
    section: SectionMetadata | None,
    column_names: list[str],
) -> None:
    """
    Fill in blank Photo, Shelf Location, Est. Linear Meters, and Shelf
    Levels from the section metadata when the row's own values are empty.

    Mutates *row_dict* in place.

    Args:
        row_dict: Row data keyed by column name.
        section: Active section metadata (may be None).
        column_names: Header column names (used to find the right keys).
    """
    if section is None:
        return

    # Build a lowercase → original-key lookup for the row dict
    lower_to_key: dict[str, str] = {k.lower(): k for k in row_dict}

    fill_map: list[tuple[list[str], object]] = [
        (["photo file name", "photo"], section.photo),
        (["shelf location"], section.shelf_location),
        (["est. linear meters"], section.est_linear_meters),
        (["shelf levels"], section.shelf_levels),
    ]

    for candidate_keys, section_value in fill_map:
        if section_value is None:
            continue
        for candidate in candidate_keys:
            actual_key = lower_to_key.get(candidate)
            if actual_key is not None:
                current = row_dict.get(actual_key)
                if current is None or str(current).strip() == "":
                    row_dict[actual_key] = section_value
                break


# ── Main extraction loop ───────────────────────────────────────────────

def _extract_rows_to_dataframe(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    column_names: list[str],
    header_row: int,
    data_start_col: int,
    merged_row_numbers: set[int],
    sections: list[SectionMetadata],
) -> tuple[pd.DataFrame, list[dict], list[SectionMetadata]]:
    """
    Walk every row in the worksheet, classify it, and build the DataFrame.

    For each row after the first header:
      - Skip if it's a merged separator row.
      - Skip if it's a repeated header row (log warning if columns differ).
      - Skip if it has fewer than MIN_CELLS_FOR_DATA_ROW non-empty cells.
      - Detect context-only rows → convert to section metadata.
      - Otherwise → add to DataFrame, applying section defaults for blanks.

    Args:
        worksheet: The openpyxl worksheet.
        column_names: Column names from the first header row.
        header_row: 1-based row number of the first header.
        data_start_col: 0-based column offset.
        merged_row_numbers: Set of rows that are merged separators.
        sections: SectionMetadata list from merged separators.

    Returns:
        (DataFrame, skipped_rows_list, context_sections_list)
    """
    data_rows: list[dict[str, object]] = []
    skipped_rows: list[dict] = []
    context_sections: list[SectionMetadata] = []

    # Keep a running pointer to the active section (from both merged and
    # context-only sources).  All sections combined, sorted by row.
    active_sections = list(sections)

    # Track the first header's column signature for consistency checks
    first_header_signature = [
        str(c).strip().lower() for c in column_names if not c.startswith("_unnamed_")
    ]

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        # ── Skip merged separator rows ────────────────────────────
        if row_idx in merged_row_numbers:
            skipped_rows.append({
                "row": row_idx,
                "reason": "merged section separator",
            })
            continue

        # ── Read cell values starting at data_start_col ───────────
        row_cells = list(
            worksheet.iter_rows(
                min_row=row_idx, max_row=row_idx, values_only=False
            )
        )[0]

        row_values: list[object] = []
        for cell in row_cells:
            if cell.column - 1 < data_start_col:
                continue
            row_values.append(cell.value)

        # Trim trailing Nones that extend past our column count
        row_values = row_values[: len(column_names)]

        # Pad if the row is shorter than the header
        while len(row_values) < len(column_names):
            row_values.append(None)

        # ── Count non-empty cells ─────────────────────────────────
        non_empty_count = sum(
            1 for v in row_values if v is not None and str(v).strip() != ""
        )

        # ── Skip rows with too few populated cells ────────────────
        if non_empty_count < MIN_CELLS_FOR_DATA_ROW:
            if non_empty_count > 0:
                skipped_rows.append({
                    "row": row_idx,
                    "reason": f"too few non-empty cells ({non_empty_count})",
                    "values": [v for v in row_values if v is not None],
                })
            # Truly empty rows are silently skipped
            continue

        # ── Detect repeated header rows ───────────────────────────
        if _is_header_row(row_values, KNOWN_HEADER_NAMES):
            # Check consistency with the first header
            current_signature = [
                str(v).strip().lower()
                for v in row_values
                if v is not None and str(v).strip().lower() in KNOWN_HEADER_NAMES
            ]
            if set(current_signature) != set(first_header_signature):
                logger.warning(
                    f"Row {row_idx}: repeated header has different columns "
                    f"than the first header at row {header_row}. "
                    f"Differences: "
                    f"added={set(current_signature) - set(first_header_signature)}, "
                    f"removed={set(first_header_signature) - set(current_signature)}"
                )
            skipped_rows.append({
                "row": row_idx,
                "reason": "repeated header row",
            })
            continue

        # ── Detect context-only rows ──────────────────────────────
        if _is_context_only_row(row_values, column_names):
            context_meta = _extract_context_metadata(
                row_values, column_names, row_idx
            )
            context_sections.append(context_meta)
            active_sections.append(context_meta)
            active_sections.sort(key=lambda s: s.start_row)
            skipped_rows.append({
                "row": row_idx,
                "reason": "context-only row (section metadata)",
            })
            continue

        # ── This is a real data row ───────────────────────────────
        row_dict: dict[str, object] = {}
        for col_name, value in zip(column_names, row_values):
            row_dict[col_name] = value

        # Apply section defaults for blank Photo/Location/etc.
        active_section = _get_section_for_row(row_idx, active_sections)
        _apply_section_defaults(row_dict, active_section, column_names)

        # Tag with source row for traceability
        row_dict["_source_row"] = row_idx

        data_rows.append(row_dict)

    raw_dataframe = pd.DataFrame(data_rows)
    return raw_dataframe, skipped_rows, context_sections
