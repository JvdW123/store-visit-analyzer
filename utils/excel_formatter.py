"""
Excel formatter — writes the output Excel workbook with three formatted sheets.

Sheet 1: "SKU Data" — full dataset with auto-filters, number formats, column
         widths, and yellow highlighting for flagged cells.
Sheet 2: "Data Quality Report" — summary stats, null counts, normalization
         log, and remaining flagged items.
Sheet 3: "Source Files" — audit trail of which files were processed.

Public API:
    format_and_save(dataframe, quality_report, source_files_info,
                    flagged_cells, output_path) → Path

See docs/SCHEMA.md — Output Excel Format for the specification.
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
import pandas as pd

from config.schema import COLUMN_TYPES, MASTER_COLUMNS
from processing.quality_checker import QualityReport

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════

_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_NORMAL_FONT = Font(size=10)
_BOLD_FONT = Font(bold=True, size=10)

# Max column width (characters) to prevent excessively wide columns
_MAX_COL_WIDTH = 50
_MIN_COL_WIDTH = 8

# Number format strings for openpyxl
_NUMBER_FORMATS: dict[str, str] = {
    "Price (Local Currency)": "#,##0.00",
    "Price (EUR)": "#,##0.00",
    "Price per Liter (EUR)": "#,##0.00",
    "Shelf Levels": "#,##0",
    "Facings": "#,##0",
    "Packaging Size (ml)": "#,##0",
    "Confidence Score": "#,##0",
    "Est. Linear Meters": "#,##0.0",
}


# ═══════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════

def format_and_save(
    dataframe: pd.DataFrame,
    quality_report: QualityReport,
    source_files_info: list[dict],
    flagged_cells: dict[tuple[int, str], str],
    output_path: Path,
) -> Path:
    """
    Write a formatted Excel workbook with three sheets.

    Args:
        dataframe: The final cleaned DataFrame to export.
        quality_report: QualityReport from quality_checker.
        source_files_info: List of dicts with keys: filename, retailer,
                          city, store_format, row_count, date_processed.
        flagged_cells: Dict mapping (row_index, column_name) to reason string
                      for cells to highlight in yellow on the SKU Data sheet.
        output_path: Path where the .xlsx file should be saved.

    Returns:
        The output_path (same as input, for convenience).
    """
    workbook = openpyxl.Workbook()

    # Sheet 1: SKU Data (rename the default sheet)
    sku_sheet = workbook.active
    sku_sheet.title = "SKU Data"
    _write_sku_data_sheet(sku_sheet, dataframe, flagged_cells)

    # Sheet 2: Data Quality Report
    quality_sheet = workbook.create_sheet("Data Quality Report")
    _write_quality_report_sheet(quality_sheet, quality_report)

    # Sheet 3: Source Files
    source_sheet = workbook.create_sheet("Source Files")
    _write_source_files_sheet(source_sheet, source_files_info)

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
    workbook.close()

    logger.info(f"Excel file saved to '{output_path}'")
    return output_path


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1: SKU Data
# ═══════════════════════════════════════════════════════════════════════════

def _write_sku_data_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    dataframe: pd.DataFrame,
    flagged_cells: dict[tuple[int, str], str],
) -> None:
    """
    Write the main SKU data sheet with headers, data, formatting, and
    flagged-cell highlighting. Adds an Issue Description column if there
    are any flagged cells.
    """
    # Determine which master columns exist in the DataFrame
    columns_to_write = [c for c in MASTER_COLUMNS if c in dataframe.columns]

    if not columns_to_write:
        columns_to_write = MASTER_COLUMNS

    # Write header row
    for col_idx, col_name in enumerate(columns_to_write, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_offset, df_idx in enumerate(dataframe.index):
        excel_row = row_offset + 2  # 1-based, header is row 1

        for col_offset, col_name in enumerate(columns_to_write):
            col_idx = col_offset + 1
            value = dataframe.at[df_idx, col_name] if col_name in dataframe.columns else None

            # Convert NaN to None for cleaner Excel output
            if pd.isna(value):
                value = None

            cell = worksheet.cell(row=excel_row, column=col_idx, value=value)
            cell.font = _NORMAL_FONT

            # Highlight flagged cells in yellow
            if (df_idx, col_name) in flagged_cells:
                cell.fill = _YELLOW_FILL

    # Add Issue Description column if there are flagged cells
    if flagged_cells:
        issue_col_idx = len(columns_to_write) + 1
        
        # Write header
        header_cell = worksheet.cell(row=1, column=issue_col_idx, value="Issue Description")
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = Alignment(horizontal="center")
        
        # Write issue descriptions for each row
        for row_offset, df_idx in enumerate(dataframe.index):
            excel_row = row_offset + 2
            
            # Collect all flagged cells for this row
            row_issues = []
            for (flagged_row_idx, flagged_col), reason in flagged_cells.items():
                if flagged_row_idx == df_idx:
                    row_issues.append(f"{flagged_col}: {reason}")
            
            # Write the combined issue description
            if row_issues:
                issue_text = " | ".join(row_issues)
                cell = worksheet.cell(row=excel_row, column=issue_col_idx, value=issue_text)
                cell.font = _NORMAL_FONT
        
        # Set column width for Issue Description column
        issue_col_letter = get_column_letter(issue_col_idx)
        worksheet.column_dimensions[issue_col_letter].width = 60

    # Apply number formats
    _apply_number_formats(worksheet, columns_to_write, len(dataframe))

    # Auto-fit column widths
    _auto_fit_column_widths(worksheet)

    # Enable auto-filter on the header row
    if columns_to_write:
        # Include Issue Description column in auto-filter if it exists
        total_columns = len(columns_to_write) + (1 if flagged_cells else 0)
        last_col_letter = get_column_letter(total_columns)
        last_row = len(dataframe) + 1  # header + data rows
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Freeze the header row
    worksheet.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2: Data Quality Report
# ═══════════════════════════════════════════════════════════════════════════

def _write_quality_report_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    report: QualityReport,
) -> None:
    """Write summary statistics, null counts, normalization log, and flagged items."""
    current_row = 1

    # ── Summary section ────────────────────────────────────────────────
    worksheet.cell(row=current_row, column=1, value="Data Quality Report").font = Font(bold=True, size=14)
    current_row += 2

    summary_items = [
        ("Total SKU Rows", report.total_rows),
        ("Files Processed", len(report.files_processed)),
        ("Data Is Clean", "Yes" if report.is_clean else "No"),
        ("Invalid Categorical Values", len(report.invalid_categoricals)),
        ("Invalid Numeric Values", len(report.invalid_numerics)),
        ("Missing Required Fields", len(report.missing_required)),
        ("Flagged Items (unresolved)", len(report.flagged_items)),
    ]

    for label, value in summary_items:
        worksheet.cell(row=current_row, column=1, value=label).font = _BOLD_FONT
        worksheet.cell(row=current_row, column=2, value=value).font = _NORMAL_FONT
        current_row += 1

    # Exchange rate
    if report.exchange_rate_used:
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Exchange Rates Used").font = _BOLD_FONT
        current_row += 1
        for currency, rate in report.exchange_rate_used.items():
            worksheet.cell(row=current_row, column=1, value=currency).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=2, value=rate).font = _NORMAL_FONT
            current_row += 1

    # ── Null counts table ──────────────────────────────────────────────
    current_row += 2
    worksheet.cell(row=current_row, column=1, value="Null Counts by Column").font = Font(bold=True, size=12)
    current_row += 1

    headers = ["Column", "Null Count", "Null %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    current_row += 1

    for col_name in MASTER_COLUMNS:
        null_count = report.null_counts.get(col_name, 0)
        null_pct = report.null_percentages.get(col_name, 0.0)
        worksheet.cell(row=current_row, column=1, value=col_name).font = _NORMAL_FONT
        worksheet.cell(row=current_row, column=2, value=null_count).font = _NORMAL_FONT
        worksheet.cell(row=current_row, column=3, value=f"{null_pct}%").font = _NORMAL_FONT
        current_row += 1

    # ── Normalization log (first 100 entries) ─────────────────────────
    if report.normalization_log:
        current_row += 2
        worksheet.cell(row=current_row, column=1, value="Normalization Log (sample)").font = Font(bold=True, size=12)
        current_row += 1

        log_headers = ["Row", "Column", "Original", "Normalized", "Method"]
        for col_idx, header in enumerate(log_headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        current_row += 1

        for entry in report.normalization_log[:100]:
            worksheet.cell(row=current_row, column=1, value=entry.get("row")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=2, value=entry.get("column")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=3, value=entry.get("original")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=4, value=entry.get("normalized")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=5, value=entry.get("method")).font = _NORMAL_FONT
            current_row += 1

    # ── Remaining flagged items ───────────────────────────────────────
    if report.flagged_items:
        current_row += 2
        worksheet.cell(row=current_row, column=1, value="Remaining Flagged Items").font = Font(bold=True, size=12)
        current_row += 1

        flag_headers = ["Row", "Column", "Original Value"]
        for col_idx, header in enumerate(flag_headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        current_row += 1

        for item in report.flagged_items:
            worksheet.cell(row=current_row, column=1, value=item.get("row_index")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=2, value=item.get("column")).font = _NORMAL_FONT
            worksheet.cell(row=current_row, column=3, value=item.get("original_value")).font = _NORMAL_FONT
            current_row += 1

    _auto_fit_column_widths(worksheet)


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 3: Source Files
# ═══════════════════════════════════════════════════════════════════════════

def _write_source_files_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    source_files: list[dict],
) -> None:
    """Write the source files audit trail table."""
    headers = ["Filename", "Retailer", "City", "Store Format", "Row Count", "Date Processed"]

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_offset, file_info in enumerate(source_files):
        excel_row = row_offset + 2
        worksheet.cell(row=excel_row, column=1, value=file_info.get("filename", "")).font = _NORMAL_FONT
        worksheet.cell(row=excel_row, column=2, value=file_info.get("retailer", "")).font = _NORMAL_FONT
        worksheet.cell(row=excel_row, column=3, value=file_info.get("city", "")).font = _NORMAL_FONT
        worksheet.cell(row=excel_row, column=4, value=file_info.get("store_format", "")).font = _NORMAL_FONT
        worksheet.cell(row=excel_row, column=5, value=file_info.get("row_count", 0)).font = _NORMAL_FONT

        date_processed = file_info.get("date_processed", datetime.now().strftime("%Y-%m-%d %H:%M"))
        worksheet.cell(row=excel_row, column=6, value=date_processed).font = _NORMAL_FONT

    # Enable auto-filter
    if source_files:
        last_row = len(source_files) + 1
        worksheet.auto_filter.ref = f"A1:F{last_row}"

    _auto_fit_column_widths(worksheet)


# ═══════════════════════════════════════════════════════════════════════════
# Formatting helpers
# ═══════════════════════════════════════════════════════════════════════════

def _apply_number_formats(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[str],
    row_count: int,
) -> None:
    """
    Apply Excel number formats to numeric columns.

    Args:
        worksheet: The worksheet to format.
        columns: List of column names in order.
        row_count: Number of data rows (excluding header).
    """
    for col_offset, col_name in enumerate(columns):
        fmt = _NUMBER_FORMATS.get(col_name)
        if fmt is None:
            continue

        col_idx = col_offset + 1
        for row_idx in range(2, row_count + 2):  # skip header row
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = fmt


def _auto_fit_column_widths(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    """
    Set column widths based on content length.

    Scans all cells in each column and sets width to the longest value,
    clamped between _MIN_COL_WIDTH and _MAX_COL_WIDTH.
    """
    for column_cells in worksheet.columns:
        max_length = _MIN_COL_WIDTH

        # Get the column letter from the first cell
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        # Clamp and add a small buffer
        adjusted_width = min(max_length + 2, _MAX_COL_WIDTH)
        worksheet.column_dimensions[col_letter].width = adjusted_width
