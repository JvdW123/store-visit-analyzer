"""
Tests for utils/excel_formatter.py

Covers: three-sheet creation, auto-filter, yellow flagged-cell highlighting,
number format application, column widths, empty DataFrame handling, and
source files sheet.
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from processing.quality_checker import QualityReport
from utils.excel_formatter import format_and_save


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_test_df(rows: int = 3) -> pd.DataFrame:
    """Build a minimal DataFrame for testing Excel output."""
    return pd.DataFrame({
        "Country": ["United Kingdom"] * rows,
        "City": ["London"] * rows,
        "Retailer": ["Tesco"] * rows,
        "Store Format": ["Large"] * rows,
        "Store Name": ["Tesco London"] * rows,
        "Brand": [f"Brand_{i}" for i in range(rows)],
        "Product Type": ["Pure Juices"] * rows,
        "Facings": [3] * rows,
        "Price (Local Currency)": [2.99] * rows,
        "Currency": ["GBP"] * rows,
        "Price (EUR)": [3.53] * rows,
        "Packaging Size (ml)": [250] * rows,
        "Price per Liter (EUR)": [14.12] * rows,
        "Confidence Score": [85] * rows,
        "Est. Linear Meters": [2.5] * rows,
    })


def _make_quality_report() -> QualityReport:
    return QualityReport(
        total_rows=3,
        rows_per_file={"file1.xlsx": 3},
        null_counts={"Brand": 0, "Flavor": 3},
        null_percentages={"Brand": 0.0, "Flavor": 100.0},
        exchange_rate_used={"GBP": 1.18},
        files_processed=["file1.xlsx"],
        is_clean=True,
    )


def _make_source_files_info() -> list[dict]:
    return [{
        "filename": "file1.xlsx",
        "retailer": "Tesco",
        "city": "London",
        "store_format": "Large",
        "row_count": 3,
        "date_processed": "2026-02-17 10:00",
    }]


def _save_and_load(
    tmp_path: Path,
    df: pd.DataFrame | None = None,
    flagged: dict | None = None,
) -> openpyxl.Workbook:
    """Save an Excel file and re-open it for inspection."""
    if df is None:
        df = _make_test_df()
    output_path = tmp_path / "test_output.xlsx"
    format_and_save(
        dataframe=df,
        quality_report=_make_quality_report(),
        source_files_info=_make_source_files_info(),
        flagged_cells=flagged or {},
        output_path=output_path,
    )
    return openpyxl.load_workbook(str(output_path))


# ═══════════════════════════════════════════════════════════════════════════
# Sheet structure
# ═══════════════════════════════════════════════════════════════════════════

class TestSheetStructure:
    def test_has_three_sheets(self, tmp_path):
        wb = _save_and_load(tmp_path)
        assert len(wb.sheetnames) == 3
        wb.close()

    def test_sheet_names(self, tmp_path):
        wb = _save_and_load(tmp_path)
        assert wb.sheetnames == ["SKU Data", "Data Quality Report", "Source Files"]
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# SKU Data sheet
# ═══════════════════════════════════════════════════════════════════════════

class TestSKUDataSheet:
    def test_auto_filter_enabled(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref != ""
        wb.close()

    def test_header_row_has_values(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]
        header_values = [cell.value for cell in ws[1] if cell.value is not None]
        assert "Country" in header_values
        assert "Brand" in header_values
        wb.close()

    def test_data_rows_written(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]
        # Row 2 should have data (row 1 is header)
        row_2_values = [cell.value for cell in ws[2]]
        non_none = [v for v in row_2_values if v is not None]
        assert len(non_none) > 0
        wb.close()

    def test_frozen_panes(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]
        assert ws.freeze_panes == "A2"
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Flagged cell highlighting
# ═══════════════════════════════════════════════════════════════════════════

class TestFlaggedHighlighting:
    def test_flagged_cell_has_yellow_fill(self, tmp_path):
        df = _make_test_df(rows=2)
        flagged = {(0, "Product Type"): "Test reason"}
        wb = _save_and_load(tmp_path, df=df, flagged=flagged)
        ws = wb["SKU Data"]

        # Find the column index for "Product Type"
        header_values = [cell.value for cell in ws[1]]
        pt_col_idx = header_values.index("Product Type") + 1  # 1-based

        # Row 2 = first data row (index 0)
        cell = ws.cell(row=2, column=pt_col_idx)
        assert cell.fill.start_color.rgb == "00FFFF00"
        wb.close()

    def test_non_flagged_cell_no_yellow(self, tmp_path):
        df = _make_test_df(rows=2)
        flagged = {(0, "Product Type"): "Test reason"}
        wb = _save_and_load(tmp_path, df=df, flagged=flagged)
        ws = wb["SKU Data"]

        # Find Brand column
        header_values = [cell.value for cell in ws[1]]
        brand_col_idx = header_values.index("Brand") + 1

        # Row 2 Brand should NOT be yellow
        cell = ws.cell(row=2, column=brand_col_idx)
        assert cell.fill.start_color.rgb != "00FFFF00" or cell.fill.fill_type is None
        wb.close()

    def test_issue_description_column_added(self, tmp_path):
        """Test that Issue Description column is added when there are flagged cells."""
        df = _make_test_df(rows=2)
        flagged = {
            (0, "Product Type"): "'Invalid Type' not in allowed values for Product Type",
            (0, "Brand"): "Brand requires review"
        }
        wb = _save_and_load(tmp_path, df=df, flagged=flagged)
        ws = wb["SKU Data"]

        # Check that Issue Description column exists in header
        header_values = [cell.value for cell in ws[1]]
        assert "Issue Description" in header_values
        
        # Find Issue Description column index
        issue_col_idx = header_values.index("Issue Description") + 1
        
        # Check that row 2 (index 0) has the combined issue description
        issue_cell = ws.cell(row=2, column=issue_col_idx)
        issue_text = issue_cell.value
        assert issue_text is not None
        assert "Product Type:" in issue_text
        assert "Brand:" in issue_text
        assert "|" in issue_text  # Separator between issues
        
        wb.close()

    def test_no_issue_description_when_no_flags(self, tmp_path):
        """Test that Issue Description column is NOT added when there are no flagged cells."""
        df = _make_test_df(rows=2)
        wb = _save_and_load(tmp_path, df=df, flagged={})
        ws = wb["SKU Data"]

        # Check that Issue Description column does NOT exist
        header_values = [cell.value for cell in ws[1]]
        assert "Issue Description" not in header_values
        
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Number formats
# ═══════════════════════════════════════════════════════════════════════════

class TestNumberFormats:
    def test_price_column_format(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]

        header_values = [cell.value for cell in ws[1]]
        if "Price (EUR)" in header_values:
            price_col_idx = header_values.index("Price (EUR)") + 1
            cell = ws.cell(row=2, column=price_col_idx)
            assert cell.number_format == "#,##0.00"
        wb.close()

    def test_integer_column_format(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]

        header_values = [cell.value for cell in ws[1]]
        if "Facings" in header_values:
            facings_col_idx = header_values.index("Facings") + 1
            cell = ws.cell(row=2, column=facings_col_idx)
            assert cell.number_format == "#,##0"
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Column widths
# ═══════════════════════════════════════════════════════════════════════════

class TestColumnWidths:
    def test_columns_have_reasonable_width(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["SKU Data"]
        for col_dim in ws.column_dimensions.values():
            if col_dim.width is not None:
                assert col_dim.width >= 8, f"Column too narrow: {col_dim.width}"
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Empty DataFrame
# ═══════════════════════════════════════════════════════════════════════════

class TestEmptyDataFrame:
    def test_empty_df_creates_valid_excel(self, tmp_path):
        df = pd.DataFrame(columns=["Country", "City", "Retailer", "Brand"])
        report = QualityReport(total_rows=0)
        output_path = tmp_path / "empty_output.xlsx"
        format_and_save(
            dataframe=df,
            quality_report=report,
            source_files_info=[],
            flagged_cells=set(),
            output_path=output_path,
        )
        wb = openpyxl.load_workbook(str(output_path))
        assert len(wb.sheetnames) == 3
        # Header should still exist
        ws = wb["SKU Data"]
        header_values = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(header_values) > 0
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Source Files sheet
# ═══════════════════════════════════════════════════════════════════════════

class TestSourceFilesSheet:
    def test_has_correct_headers(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["Source Files"]
        header_values = [cell.value for cell in ws[1]]
        assert "Filename" in header_values
        assert "Retailer" in header_values
        assert "Row Count" in header_values
        assert "Date Processed" in header_values
        wb.close()

    def test_data_row_written(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["Source Files"]
        # Row 2 should have "file1.xlsx"
        assert ws.cell(row=2, column=1).value == "file1.xlsx"
        wb.close()

    def test_auto_filter_on_source_files(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["Source Files"]
        assert ws.auto_filter.ref is not None
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# Quality Report sheet
# ═══════════════════════════════════════════════════════════════════════════

class TestQualityReportSheet:
    def test_title_present(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["Data Quality Report"]
        assert ws.cell(row=1, column=1).value == "Data Quality Report"
        wb.close()

    def test_exchange_rate_present(self, tmp_path):
        wb = _save_and_load(tmp_path)
        ws = wb["Data Quality Report"]
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v is not None])
        joined = " ".join(all_values)
        assert "GBP" in joined
        wb.close()
